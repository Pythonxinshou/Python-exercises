from openpyxl import load_workbook


"""plsql自带导出数据生成insertSQL文件"""
# 设置路径
path = 'C:/Users/zsj/PycharmProjects/Python-exercises/my_trials/sql.xlsx'
# 打开文件
wb = load_workbook(path)

# 获得所有sheet的名称
# print(wb.get_sheet_names())
# 根据sheet名字获得sheet
a_sheet = wb.get_sheet_by_name('Sheet1')
# 获得sheet名
# print(a_sheet.title)
# 获得当前正在显示的sheet, 也可以用wb.get_active_sheet()
data_sheet = wb.active

# 获得最大行和最大列
# print(data_sheet.max_row)
# print(data_sheet.max_column)

# 创建三个列表包含字段名、值以及值为空的索引位列表
key_list, value_list, num_list = [], [], []

for i in range(data_sheet.max_row):
    key_list.append(data_sheet.cell(row=i+1, column=1).value)
for j in range(data_sheet.max_row):
    if not data_sheet.cell(j+1, 2).value:
        num_list.append(j)
    else:
        value_list.append(data_sheet.cell(row=j+1, column=2).value)
print(key_list, value_list, num_list, '\n')

# 提出值为空的字段
for i in range(-1, -len(num_list)-1, -1):
    del key_list[num_list[i]]
a, b = '', ''
for i in range(len(key_list)):
    if i != len(key_list)-1:
        a += "'{}',".format(key_list[i])
        b += "'{}',".format(value_list[i])
    else:
        a += "'{}'".format(key_list[i])
        b += "'{}'".format(value_list[i])

c = "INSERT INTO table_name (" + a + ") VALUES (" + b + ")"

print(c)
